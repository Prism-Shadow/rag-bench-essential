from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest import mock

from openpyxl import Workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS_DIR = REPO_ROOT / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from scoring.report import run_validator
from scoring.rule_checks import check_banker_workbook


TRUTH_DIR = REPO_ROOT / "cases/bankertoolbench_cake_lbo_sensitivity_hard/truth"
EXPECTED = json.loads((TRUTH_DIR / "expected.json").read_text(encoding="utf-8"))


def build_workbook(path: Path, *, variant_titles: bool, valid_grid: bool) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = EXPECTED["model_sheet"]
    sheet[EXPECTED["base_irr_cell"]] = "=0.25"
    for index, table in enumerate(EXPECTED["tables"]):
        title_row = 10 + index * 10
        header_row = title_row + 1
        grid_row = title_row + 2
        grid_col = 3
        sheet.cell(title_row, 2).value = (
            f"IRR sensitivity: {table['id']}" if variant_titles else table["title"]
        )
        for offset, value in enumerate(table["x_values"]):
            sheet.cell(header_row, grid_col + offset).value = value
        for offset, value in enumerate(table["y_values"]):
            sheet.cell(grid_row + offset, grid_col - 1).value = value
        for row_offset in range(5):
            for col_offset in range(5):
                sheet.cell(grid_row + row_offset, grid_col + col_offset).value = (
                    "=0.25" if valid_grid else 5.0
                )
    workbook.save(path)


def write_pptx(path: Path) -> None:
    namespace = "http://schemas.openxmlformats.org/drawingml/2006/main"
    tokens = EXPECTED["ppt_required_tokens"]
    with zipfile.ZipFile(path, "w") as archive:
        for index in (1, 2):
            text = " ".join(tokens[index - 1 :: 2])
            archive.writestr(
                f"ppt/slides/slide{index}.xml",
                f'<root xmlns:a="{namespace}"><a:t>{text}</a:t></root>',
            )


class BankerScoringTest(unittest.TestCase):
    def test_valid_structure_does_not_require_exact_titles(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "variant.xlsx"
            build_workbook(path, variant_titles=True, valid_grid=True)
            result = check_banker_workbook(path, EXPECTED)
        self.assertTrue(result.passed, result.details)
        self.assertTrue(
            all(not table["candidate"]["title_match"] for table in result.details["tables"])
        )

    def test_axes_and_titles_without_irr_grid_fail(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "bad-grid.xlsx"
            build_workbook(path, variant_titles=False, valid_grid=False)
            result = check_banker_workbook(path, EXPECTED)
        self.assertFalse(result.passed)
        self.assertTrue(all(table["y_axis_found"] for table in result.details["tables"]))
        self.assertTrue(all(not table["passed"] for table in result.details["tables"]))

    def test_external_validator_uses_the_shared_workbook_check(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            workspace = Path(directory)
            build_workbook(
                workspace / EXPECTED["required_outputs"][0],
                variant_titles=True,
                valid_grid=True,
            )
            write_pptx(workspace / EXPECTED["required_outputs"][1])
            (workspace / EXPECTED["required_outputs"][2]).write_bytes(
                b"%PDF-1.4\n" + b"0" * 1200
            )
            env = os.environ.copy()
            env["BENCH_TRUTH_DIR"] = str(TRUTH_DIR)
            result = subprocess.run(
                [sys.executable, str(TRUTH_DIR / "validate.py")],
                cwd=workspace,
                env=env,
                text=True,
                capture_output=True,
                check=False,
            )
        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_validator_runner_uses_current_interpreter(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            truth = root / "truth"
            workspace = root / "workspace"
            truth.mkdir()
            workspace.mkdir()
            (truth / "validate.py").write_text("raise SystemExit(0)\n", encoding="utf-8")
            completed = subprocess.CompletedProcess([], 0, "", "")
            with mock.patch("scoring.report.subprocess.run", return_value=completed) as run:
                run_validator(workspace, truth)
        command = run.call_args.args[0]
        self.assertEqual(command[0], sys.executable)


if __name__ == "__main__":
    unittest.main()
