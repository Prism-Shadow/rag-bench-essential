from __future__ import annotations

import csv
import json
import math
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


@dataclass
class CheckResult:
    passed: bool
    message: str
    details: dict[str, Any] | None = None


@dataclass
class ScoreContext:
    workspace: Path
    truth_dir: Path
    expected: dict[str, Any]
    validator_exit: int | None


def run_check(check: dict[str, Any], ctx: ScoreContext) -> CheckResult:
    check_type = check.get("type")
    if check_type == "file_exists":
        return file_exists(ctx.workspace / check["path"])
    if check_type == "json_parse":
        return json_parse(ctx.workspace / check["path"])
    if check_type == "validator_exit":
        return validator_exit(ctx.validator_exit, check.get("expected", 0))
    if check_type == "required_outputs_present":
        return required_outputs_present(ctx.workspace, _required_outputs(ctx.expected))
    if check_type == "dabstep_fee_ids":
        return dabstep_fee_ids(ctx)
    if check_type == "numeric_answer":
        return numeric_answer(ctx)
    if check_type == "numeric_vector_answer":
        return numeric_vector_answer(ctx)
    if check_type == "numeric_vector_element":
        return numeric_vector_element(ctx, int(check.get("index", -1)))
    if check_type == "dci_answer":
        return dci_answer(ctx)
    if check_type == "dci_evidence":
        return dci_evidence(ctx)
    if check_type == "dci_intermediates":
        return dci_intermediates(ctx)
    if check_type == "dvworld_chart_spec":
        return dvworld_chart_spec(ctx)
    if check_type == "dvworld_sheet_answers":
        return dvworld_sheet_answers(ctx)
    if check_type == "dvworld_sheet_derived_data":
        return dvworld_sheet_derived_data(ctx)
    if check_type == "dvworld_sheet_chart_spec":
        return dvworld_sheet_chart_spec(ctx)
    if check_type == "png_valid":
        return png_valid(ctx.workspace / check["path"])
    if check_type == "taobao_csv":
        return taobao_csv(ctx)
    if check_type == "taobao_json_rules":
        return taobao_json_rules(ctx)
    if check_type == "taobao_markdown":
        return taobao_markdown(ctx)
    if check_type == "banker_workbook":
        return banker_workbook(ctx)
    if check_type == "banker_ppt":
        return banker_ppt(ctx)
    if check_type == "banker_pdf":
        return banker_pdf(ctx)
    if check_type == "harvey_docx_delivery":
        return harvey_docx_delivery(ctx)
    if check_type == "harvey_issue_coverage":
        return harvey_issue_coverage(ctx)
    if check_type == "harvey_source_binding":
        return harvey_source_binding(ctx)
    if check_type == "harvey_memo_controls":
        return harvey_memo_controls(ctx)
    if check_type == "finlong_universe":
        return finlong_universe(ctx)
    if check_type == "finlong_ranking":
        return finlong_ranking(ctx)
    if check_type == "finlong_figures":
        return finlong_figures(ctx)
    if check_type == "medagent_actions":
        return medagent_actions(ctx)
    if check_type == "medagent_report":
        return medagent_report(ctx)
    if check_type == "prepbench_by_tier":
        return prepbench_by_tier(ctx)
    if check_type == "spider_part_output":
        return spider_part_output(ctx, str(check.get("part", "")))
    if check_type == "spider_part_outputs":
        return spider_part_outputs(ctx)
    if check_type == "spreadsheet_workbooks":
        return spreadsheet_workbooks(ctx)
    if check_type == "spreadsheet_workbook_variant":
        return spreadsheet_workbooks(ctx, str(check.get("variant", "")))
    return CheckResult(False, f"unknown check type: {check_type!r}")


def _required_outputs(expected: dict[str, Any]) -> list[str]:
    if "required_outputs" in expected:
        paths: list[str] = []
        for item in expected["required_outputs"]:
            if isinstance(item, str):
                paths.append(item)
            elif isinstance(item, dict) and item.get("output"):
                paths.append(str(item["output"]))
        return paths
    if "required_files" in expected:
        return list(expected["required_files"])
    if "required_output" in expected:
        return [expected["required_output"]]
    if "output_file" in expected:
        return [expected["output_file"]]
    if "outputs" in expected and isinstance(expected["outputs"], dict):
        return list(expected["outputs"].values())
    return ["answers.json"]


def file_exists(path: Path) -> CheckResult:
    return CheckResult(path.exists(), f"{path.name} exists" if path.exists() else f"missing {path}")


def json_parse(path: Path) -> CheckResult:
    try:
        json.loads(path.read_text(encoding="utf-8"))
        return CheckResult(True, f"{path.name} parses as JSON")
    except Exception as exc:
        return CheckResult(False, f"{path} is not parseable JSON: {exc}")


def validator_exit(actual: int | None, expected: int) -> CheckResult:
    return CheckResult(actual == expected, f"validator_exit expected {expected}, got {actual}")


def required_outputs_present(workspace: Path, paths: list[str]) -> CheckResult:
    missing = [path for path in paths if not (workspace / path).exists()]
    return CheckResult(not missing, "all required outputs present" if not missing else "missing outputs", {"missing": missing})


def _load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _as_id_set(value: Any) -> set[str]:
    if value is None:
        return set()
    if isinstance(value, (list, tuple)):
        value = ", ".join(str(v) for v in value)
    text = str(value).strip()
    if text == "" or text.lower() in {"not applicable", "n/a", "none"}:
        return set()
    return {token.strip() for token in re.split(r"[,\s]+", text) if token.strip()}


def dabstep_fee_ids(ctx: ScoreContext) -> CheckResult:
    try:
        answers = _load_json(ctx.workspace / "answers.json")
    except Exception as exc:
        return CheckResult(False, f"answers.json unreadable: {exc}")
    got = _as_id_set(answers.get("answer"))
    expected = _as_id_set(ctx.expected.get("answer"))
    passed = got == expected
    return CheckResult(
        passed,
        "fee ID set matches" if passed else "fee ID set mismatch",
        {"missing": sorted(expected - got), "extra": sorted(got - expected), "got_count": len(got), "expected_count": len(expected)},
    )


def _answer_value(answers: Any) -> Any:
    if isinstance(answers, dict):
        return answers.get("answer")
    return answers


def _number_from_answer(value: Any) -> float | None:
    if isinstance(value, (list, tuple)):
        value = value[0] if value else None
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().lower()
    if text in {"", "not applicable", "n/a", "none"}:
        return None
    text = text.replace(",", "").replace("$", "").replace("%", "").replace("eur", "").strip()
    match = re.search(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", text)
    return float(match.group(0)) if match else None


def numeric_answer(ctx: ScoreContext) -> CheckResult:
    try:
        answers = _load_json(ctx.workspace / "answers.json")
    except Exception as exc:
        return CheckResult(False, f"answers.json unreadable: {exc}")
    raw_expected = ctx.expected.get("gold_numeric", ctx.expected.get("answer"))
    got = _number_from_answer(_answer_value(answers))
    expected = _number_from_answer(raw_expected)
    if expected is None or got is None:
        return CheckResult(False, "numeric answer missing", {"expected": raw_expected, "got": _answer_value(answers)})
    if "abs_tolerance" in ctx.expected:
        tol = float(ctx.expected["abs_tolerance"])
    elif "rel_tolerance" in ctx.expected:
        tol = max(abs(expected) * float(ctx.expected["rel_tolerance"]), 1e-9)
    else:
        tol = float(ctx.expected.get("tolerance", 1e-9))
    passed = math.isfinite(got) and abs(got - expected) <= tol
    return CheckResult(passed, "numeric answer matches" if passed else "numeric answer mismatch", {"expected": expected, "got": got, "tolerance": tol})


def _numeric_vector(value: Any) -> list[float] | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return [float(value)]
    if isinstance(value, str):
        nums = re.findall(r"[-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?", value.replace(",", ""))
        return [float(item) for item in nums] if nums else None
    if isinstance(value, (list, tuple)):
        out: list[float] = []
        for item in value:
            sub = _numeric_vector(item)
            if sub is None:
                return None
            out.extend(sub)
        return out
    return None


def numeric_vector_answer(ctx: ScoreContext) -> CheckResult:
    try:
        answers = _load_json(ctx.workspace / "answers.json")
    except Exception as exc:
        return CheckResult(False, f"answers.json unreadable: {exc}")
    got = _numeric_vector(_answer_value(answers))
    expected = _numeric_vector(ctx.expected.get("answer"))
    if expected is None or got is None:
        return CheckResult(False, "numeric vector missing", {"expected": ctx.expected.get("answer"), "got": _answer_value(answers)})
    if len(got) != len(expected):
        return CheckResult(False, "numeric vector length mismatch", {"expected_len": len(expected), "got_len": len(got), "expected": expected, "got": got})
    rel = float(ctx.expected.get("tolerance", 0.05))
    bad = []
    for idx, (actual, exp) in enumerate(zip(got, expected)):
        tol = max(abs(exp) * rel, 1e-9)
        if abs(actual - exp) > tol:
            bad.append({"idx": idx, "expected": exp, "got": actual, "tol": tol})
    return CheckResult(not bad, "numeric vector matches" if not bad else "numeric vector mismatch", {"bad": bad, "expected": expected, "got": got})


def numeric_vector_element(ctx: ScoreContext, index: int) -> CheckResult:
    try:
        answers = _load_json(ctx.workspace / "answers.json")
    except Exception as exc:
        return CheckResult(False, f"answers.json unreadable: {exc}")
    got = _numeric_vector(_answer_value(answers))
    expected = _numeric_vector(ctx.expected.get("answer"))
    if expected is None or got is None:
        return CheckResult(False, "numeric vector missing")
    if len(got) != len(expected):
        return CheckResult(
            False,
            "numeric vector length mismatch",
            {"expected_len": len(expected), "got_len": len(got)},
        )
    if index < 0 or index >= len(expected):
        return CheckResult(False, f"numeric vector index out of range: {index}")
    tolerance = float(ctx.expected.get("abs_tolerance", 1.0))
    passed = math.isfinite(got[index]) and abs(got[index] - expected[index]) <= tolerance
    return CheckResult(
        passed,
        f"numeric vector element {index} matches" if passed else f"numeric vector element {index} mismatch",
        {"index": index, "expected": expected[index], "got": got[index], "abs_tolerance": tolerance},
    )


def _dci_normalize(value: Any) -> str:
    if isinstance(value, list):
        value = " ".join(str(item) for item in value)
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s_\\-]+", " ", text)
    text = re.sub(r"[^a-z0-9 ]+", "", text)
    return re.sub(r"\s+", " ", text).strip()


def _dci_normalize_ref(value: Any) -> str:
    text = str(value or "").strip().lower().replace("\\", "/")
    text = re.sub(r"[^a-z0-9/.: -]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _flatten_text(value: Any) -> str:
    if isinstance(value, dict):
        return "\n".join(f"{key}: {_flatten_text(item)}" for key, item in value.items())
    if isinstance(value, list):
        return "\n".join(_flatten_text(item) for item in value)
    return str(value or "")


def _read_text_if_exists(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except OSError:
        return ""


def dci_answer(ctx: ScoreContext) -> CheckResult:
    try:
        answers = _load_json(ctx.workspace / "answers.json")
    except Exception as exc:
        return CheckResult(False, f"answers.json unreadable: {exc}")
    answer_value = _answer_value(answers)
    aliases = [ctx.expected["answer"]["exact"], *ctx.expected["answer"].get("aliases", [])]
    passed = _dci_normalize(answer_value) in {_dci_normalize(alias) for alias in aliases}
    return CheckResult(passed, "DCI answer matches" if passed else "DCI answer mismatch", {"expected": aliases, "got": answer_value})


def _dci_doc_referenced(text: str, doc: dict[str, Any]) -> bool:
    lower = text.lower()
    norm = _dci_normalize_ref(text)
    for key in ("docid", "url", "path"):
        value = str(doc.get(key) or "").strip()
        if value and value.lower() in lower:
            return True
    title = str(doc.get("title") or "").strip()
    if title and _dci_normalize_ref(title) in norm:
        return True
    path = str(doc.get("path") or "").strip()
    return bool(path and _dci_normalize_ref(Path(path).name) in norm)


def _dci_ref_source_text(ctx: ScoreContext, evidence_text: str, ref: dict[str, Any]) -> str:
    parts = [evidence_text]
    for key in ("path", "file", "source_path"):
        value = ref.get(key)
        if not value:
            continue
        candidate = Path(str(value))
        if not candidate.is_absolute():
            candidate = ctx.workspace / candidate
        if candidate.exists() and candidate.is_file():
            parts.append(_read_text_if_exists(candidate))
    return "\n".join(parts)


def _dci_docs_by_role(expected: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        str(doc.get("role")): doc
        for doc in expected.get("required_evidence_docs", [])
        if doc.get("role")
    }


def _dci_term_groups_hit(text: str, groups: list[list[str]]) -> tuple[bool, list[list[str]]]:
    lower = text.lower()
    missing = [
        group
        for group in groups
        if not any(str(term).lower() in lower for term in group)
    ]
    return not missing, missing


def _dci_matching_refs_for_group(
    source_refs: list[Any],
    expected: dict[str, Any],
    group: dict[str, Any],
) -> list[dict[str, Any]]:
    role_docs = _dci_docs_by_role(expected)
    accepted_roles = group.get("accepted_doc_roles") or [group.get("role")]
    accepted_docs = [
        role_docs[role]
        for role in accepted_roles
        if role in role_docs
    ]
    matches: list[dict[str, Any]] = []
    for ref in source_refs:
        if not isinstance(ref, dict):
            continue
        flat = _flatten_text(ref)
        if any(_dci_doc_referenced(flat, doc) for doc in accepted_docs):
            matches.append(ref)
    return matches


def _dci_source_supports_constraint(ref: Any, key: str, aliases: dict[str, list[str]]) -> bool:
    if not isinstance(ref, dict):
        return False
    supports = ref.get("supports", [])
    if not isinstance(supports, list):
        supports = [supports]
    text = _dci_normalize_ref("\n".join(str(item) for item in supports))
    accepted = [key, *aliases.get(key, [])]
    return any(_dci_normalize_ref(item) in text for item in accepted)


def dci_evidence(ctx: ScoreContext) -> CheckResult:
    try:
        evidence = _load_json(ctx.workspace / "evidence.json")
    except Exception as exc:
        return CheckResult(False, f"evidence.json unreadable: {exc}")
    evidence_text = _flatten_text(evidence)
    report_text = _read_text_if_exists(ctx.workspace / "report.md")
    combined_text = "\n".join([evidence_text, report_text])
    source_refs = evidence.get("sources", []) if isinstance(evidence, dict) else []
    if not isinstance(source_refs, list):
        source_refs = []
    misses: list[str] = []
    missing_term_groups: dict[str, list[list[str]]] = {}
    groups = ctx.expected.get("required_evidence_groups")
    if groups:
        for group in groups:
            matching_refs = _dci_matching_refs_for_group(source_refs, ctx.expected, group)
            haystack = "\n".join(_dci_ref_source_text(ctx, evidence_text, ref) for ref in matching_refs)
            term_ok, missing_groups = _dci_term_groups_hit(haystack, group.get("must_include_any_groups", []))
            if not matching_refs or not term_ok:
                misses.append(group["id"])
                if missing_groups:
                    missing_term_groups[group["id"]] = missing_groups
        return CheckResult(
            not misses,
            "DCI evidence binding passes" if not misses else "DCI evidence binding misses required groups",
            {"missing_groups": misses, "missing_term_groups": missing_term_groups},
        )
    for doc in ctx.expected["required_evidence_docs"]:
        matching_refs = [ref for ref in source_refs if _dci_doc_referenced(_flatten_text(ref), doc)]
        haystack = "\n".join(_dci_ref_source_text(ctx, evidence_text, ref) for ref in matching_refs) or combined_text
        phrase_hit = any(term.lower() in haystack.lower() for term in doc.get("must_include_any", []))
        ref_hit = _dci_doc_referenced(combined_text, doc)
        if not (ref_hit and phrase_hit):
            misses.append(doc["role"])
    return CheckResult(not misses, "DCI evidence binding passes" if not misses else "DCI evidence binding misses required docs", {"missing_roles": misses})


def dci_intermediates(ctx: ScoreContext) -> CheckResult:
    try:
        evidence = _load_json(ctx.workspace / "evidence.json")
    except Exception as exc:
        return CheckResult(False, f"evidence.json unreadable: {exc}")
    combined = "\n".join([_flatten_text(evidence), _read_text_if_exists(ctx.workspace / "report.md")]).lower()
    missing_terms = [term for term in ctx.expected["required_intermediate_terms"] if term.lower() not in combined]
    checklist = evidence.get("constraint_checklist", {}) if isinstance(evidence, dict) else {}
    if not isinstance(checklist, dict):
        checklist = {}
    support_aliases = ctx.expected.get("constraint_support_aliases", {})
    source_refs = evidence.get("sources", []) if isinstance(evidence, dict) else []
    if not isinstance(source_refs, list):
        source_refs = []
    missing_checklist = []
    missing_support = []
    for key in ctx.expected.get("required_checklist_true", []):
        if checklist.get(key) is not True:
            missing_checklist.append(key)
        if not any(_dci_source_supports_constraint(ref, key, support_aliases) for ref in source_refs):
            missing_support.append(key)
    passed = not missing_terms and not missing_checklist and not missing_support
    return CheckResult(
        passed,
        "DCI intermediate chain passes" if passed else "DCI intermediate chain incomplete",
        {"missing_terms": missing_terms, "missing_checklist": missing_checklist, "missing_support": missing_support},
    )


def _float_close(actual: Any, expected: Any, tol: float) -> bool:
    try:
        value = float(actual)
    except Exception:
        return False
    return not math.isnan(value) and abs(value - float(expected)) <= tol


def dvworld_chart_spec(ctx: ScoreContext) -> CheckResult:
    try:
        spec = _load_json(ctx.workspace / "chart_spec.json")
    except Exception as exc:
        return CheckResult(False, f"chart_spec.json unreadable: {exc}")
    errors: list[str] = []
    chart = ctx.expected["chart"]
    inter = ctx.expected["key_intermediates"]
    if spec.get("source_file") != inter["source_file"]:
        errors.append("source_file")
    if spec.get("year") != inter["year"]:
        errors.append("year")
    if spec.get("title") != chart["title"]:
        errors.append("title")
    if spec.get("mark") != "network":
        errors.append("mark")
    layout = spec.get("layout", {})
    for key, value in chart["layout"].items():
        if layout.get(key) != value:
            errors.append(f"layout.{key}")
    if layout.get("node_order") != inter["node_order"]:
        errors.append("layout.node_order")
    nodes = spec.get("nodes", {})
    if nodes.get("values") != inter["node_order"]:
        errors.append("nodes.values")
    if nodes.get("fill") != chart["node_fill"]:
        errors.append("nodes.fill")
    if nodes.get("stroke") != chart["node_stroke"]:
        errors.append("nodes.stroke")
    edges = spec.get("edges", {})
    edge_expect = {
        "source": "source",
        "target": "target",
        "weight": "corr",
        "threshold_abs_corr": inter["threshold_abs_corr"],
        "undirected": True,
        "self_loops": False,
    }
    for key, value in edge_expect.items():
        if edges.get(key) != value:
            errors.append(f"edges.{key}")
    rows = edges.get("values")
    if not isinstance(rows, list):
        errors.append("edges.values")
        rows = []
    gold_rows = ctx.expected["edges"]
    if len(rows) != len(gold_rows):
        errors.append(f"edges.values count expected {len(gold_rows)}, got {len(rows)}")
    tol = float(ctx.expected["tolerances"]["float_abs"])
    for idx, gold in enumerate(gold_rows[: len(rows)]):
        row = rows[idx]
        if not isinstance(row, dict):
            errors.append(f"edges.values[{idx}] object")
            continue
        for key in ("source", "target", "sign", "strength_bin", "edge_color_hex"):
            if row.get(key) != gold[key]:
                errors.append(f"edges.values[{idx}].{key}")
        for key in ("corr", "abs_corr", "edge_width"):
            if not _float_close(row.get(key), gold[key], tol):
                errors.append(f"edges.values[{idx}].{key}")
    encoding = spec.get("edge_encoding", {})
    color = encoding.get("color", {})
    if color.get("field") != "sign":
        errors.append("edge_encoding.color.field")
    if color.get("positive") != chart["edge_color_positive"]:
        errors.append("edge_encoding.color.positive")
    if color.get("negative") != chart["edge_color_negative"]:
        errors.append("edge_encoding.color.negative")
    width = encoding.get("width", {})
    if width.get("field") != "strength_bin":
        errors.append("edge_encoding.width.field")
    for bin_name, expected_width in chart["width_by_bin"].items():
        if not _float_close(width.get(bin_name), expected_width, tol):
            errors.append(f"edge_encoding.width.{bin_name}")
    return CheckResult(not errors, "chart_spec.json matches" if not errors else "chart_spec.json mismatch", {"errors": errors})


def png_valid(path: Path) -> CheckResult:
    try:
        signature = path.read_bytes()[:8]
    except Exception as exc:
        return CheckResult(False, f"PNG unreadable: {exc}")
    return CheckResult(signature == b"\x89PNG\r\n\x1a\n", "valid PNG header" if signature == b"\x89PNG\r\n\x1a\n" else "invalid PNG header")


def dvworld_sheet_answers(ctx: ScoreContext) -> CheckResult:
    try:
        answers = _load_json(ctx.workspace / "answers.json")
    except Exception as exc:
        return CheckResult(False, f"answers.json unreadable: {exc}")
    chart = ctx.expected["chart"]
    inter = ctx.expected["key_intermediates"]
    required = {
        "case_id": ctx.expected["case_id"],
        "chart_type": chart["chart_type"],
        "source_file": inter["source_file"],
        "sheet": inter["sheet"],
        "bar_metric": inter["bar_metric"],
        "line_metric": inter["line_metric"],
        "row_count": chart["row_count"],
        "regions": inter["regions"],
        "soil_types": inter["soil_types"],
        "secondary_axis_for_line": chart["secondary_axis_for_line"],
        "bar_value_labels": chart["bar_value_labels"],
    }
    mismatches = {key: {"expected": value, "actual": answers.get(key)} for key, value in required.items() if answers.get(key) != value}
    return CheckResult(not mismatches, "DVWorld sheet answers match" if not mismatches else "DVWorld sheet answers mismatch", {"mismatches": mismatches})


def dvworld_sheet_derived_data(ctx: ScoreContext) -> CheckResult:
    try:
        with (ctx.workspace / "derived_data.csv").open(newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
            fieldnames = reader.fieldnames
    except Exception as exc:
        return CheckResult(False, f"derived_data.csv unreadable: {exc}")
    errors: list[str] = []
    if fieldnames != ctx.expected["csv_columns"]:
        errors.append("columns mismatch")
    gold_rows = ctx.expected["derived_data"]
    if len(rows) != len(gold_rows):
        errors.append(f"row count expected {len(gold_rows)}, got {len(rows)}")
    tol = ctx.expected["tolerances"]["float_abs"]
    for idx, gold in enumerate(gold_rows[: len(rows)]):
        row = rows[idx]
        if row.get("Region") != gold["Region"]:
            errors.append(f"row {idx + 1} Region")
        for key in ("Clay", "Loam", "Sandy", "Climate_Ratio", "Rainfall_mean", "Temperature_mean"):
            if not _float_close(row.get(key), gold[key], tol):
                errors.append(f"row {idx + 1} {key}")
    return CheckResult(not errors, "derived_data.csv matches" if not errors else "derived_data.csv mismatch", {"errors": errors[:20], "error_count": len(errors)})


def _chart_layer_by_mark(spec: dict[str, Any], mark: str) -> dict[str, Any]:
    for layer in spec.get("layers", []):
        if isinstance(layer, dict) and layer.get("mark") == mark:
            return layer
    return {}


def dvworld_sheet_chart_spec(ctx: ScoreContext) -> CheckResult:
    try:
        spec = _load_json(ctx.workspace / "chart_spec.json")
    except Exception as exc:
        return CheckResult(False, f"chart_spec.json unreadable: {exc}")
    errors: list[str] = []
    chart = ctx.expected["chart"]
    inter = ctx.expected["key_intermediates"]
    if spec.get("title") != chart["title"]:
        errors.append("title")
    if spec.get("data") != "derived_data.csv":
        errors.append("data")
    x_spec = spec.get("x", {})
    if x_spec.get("field") != "Region":
        errors.append("x.field")
    if x_spec.get("order") != inter["regions"]:
        errors.append("x.order")
    bar = _chart_layer_by_mark(spec, "bar")
    if bar.get("series_fields") != inter["soil_types"]:
        errors.append("bar.series_fields")
    if bar.get("y_axis") != "primary":
        errors.append("bar.y_axis")
    if bar.get("data_labels") is not True:
        errors.append("bar.data_labels")
    line = _chart_layer_by_mark(spec, "line")
    if line.get("field") != "Climate_Ratio":
        errors.append("line.field")
    if line.get("y_axis") != "secondary":
        errors.append("line.y_axis")
    if spec.get("legend") != chart["legend"]:
        errors.append("legend")
    if spec.get("axis_titles") != chart["axis_titles"]:
        errors.append("axis_titles")
    return CheckResult(not errors, "DVWorld sheet chart spec matches" if not errors else "DVWorld sheet chart spec mismatch", {"errors": errors})


ROLE_ALIASES = {
    "管理员": "admin",
    "admin": "admin",
    "运营": "operation",
    "operation": "operation",
    "美工": "designer",
    "设计": "designer",
    "designer": "designer",
    "客服": "service",
    "service": "service",
    "仓储": "stock",
    "库存": "stock",
    "stock": "stock",
}

MODULE_ALIASES = {
    "activity_apply": "activity_apply",
    "活动报名": "activity_apply",
    "活动配置": "activity_apply",
    "item_upshelf": "item_upshelf",
    "商品上下架": "item_upshelf",
    "上下架": "item_upshelf",
    "price_update": "price_update",
    "价格修改": "price_update",
    "价格": "price_update",
    "material_edit": "material_edit",
    "素材编辑": "material_edit",
    "素材": "material_edit",
    "inventory_data_view": "inventory_data_view",
    "库存数据查看": "inventory_data_view",
    "库存": "inventory_data_view",
    "order_data_view": "order_data_view",
    "订单数据查看": "order_data_view",
    "订单": "order_data_view",
    "活动数据查看": "order_data_view",
    "data_export": "data_export",
    "数据导出": "data_export",
    "导出": "data_export",
    "log_view": "log_view",
    "日志查看": "log_view",
    "操作日志": "log_view",
    "user_manager": "user_manager",
    "用户管理": "user_manager",
}


def _normalize_role(value: str) -> str:
    raw = str(value or "").strip()
    low = raw.lower()
    if low in ROLE_ALIASES:
        return ROLE_ALIASES[low]
    for token, role in ROLE_ALIASES.items():
        if token and token in raw:
            return role
    return low


def _normalize_module(value: str) -> str:
    raw = str(value or "").strip()
    low = raw.lower()
    if low in MODULE_ALIASES:
        return MODULE_ALIASES[low]
    for token, module in MODULE_ALIASES.items():
        if token and token in raw:
            return module
    return low


def _parse_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    if text in {"true", "1", "yes", "y", "是", "可", "允许", "read", "write"}:
        return True
    if text in {"false", "0", "no", "n", "否", "不可", "禁止", "deny", ""}:
        return False
    return None


def taobao_csv(ctx: ScoreContext) -> CheckResult:
    try:
        rows = list(csv.DictReader((ctx.workspace / "淘宝活动/权限配置表.csv").open("r", encoding="utf-8-sig", newline="")))
    except Exception as exc:
        return CheckResult(False, f"CSV unreadable: {exc}")
    required_cols = {"role", "module", "read", "write", "download", "basis"}
    cols = set(rows[0].keys()) if rows else set()
    if not required_cols.issubset(cols):
        return CheckResult(False, "CSV columns missing", {"expected": sorted(required_cols), "actual": sorted(cols)})
    matrix: dict[tuple[str, str], dict[str, bool | None]] = {}
    for row in rows:
        matrix[(_normalize_role(row.get("role", "")), _normalize_module(row.get("module", "")))] = {
            "read": _parse_bool(row.get("read")),
            "write": _parse_bool(row.get("write")),
            "download": _parse_bool(row.get("download")),
        }
    total = matched = 0
    mismatches: list[str] = []
    for role, modules in ctx.expected["permissions"].items():
        for module, exp_perm in modules.items():
            total += 1
            got = matrix.get((role, module))
            if got is None:
                mismatches.append(f"missing {role}.{module}")
                continue
            keys = ["read", "write"] + (["download"] if module == "data_export" else [])
            if all(got.get(key) is exp_perm.get(key) for key in keys):
                matched += 1
            else:
                mismatches.append(f"{role}.{module}")
    return CheckResult(not mismatches, "permission CSV matches" if not mismatches else "permission CSV mismatch", {"matched": matched, "total": total, "mismatches": mismatches[:20]})


def _normalize_role_list(values: Any) -> list[str]:
    if not isinstance(values, list):
        return []
    return sorted({_normalize_role(v) for v in values})


def taobao_json_rules(ctx: ScoreContext) -> CheckResult:
    try:
        rules = _load_json(ctx.workspace / "淘宝活动/权限校验规则.json")
    except Exception as exc:
        return CheckResult(False, f"JSON rules unreadable: {exc}")
    errors: list[str] = []
    if rules.get("default_deny") is not True:
        errors.append("default_deny")
    sensitive = rules.get("sensitive_operations", {})
    if not isinstance(sensitive, dict):
        errors.append("sensitive_operations object")
    else:
        for module, exp_roles in ctx.expected["sensitive_operations"].items():
            if _normalize_role_list(sensitive.get(module)) != sorted(exp_roles):
                errors.append(f"sensitive_operations.{module}")
    return CheckResult(not errors, "JSON rules match" if not errors else "JSON rules mismatch", {"errors": errors})


def taobao_markdown(ctx: ScoreContext) -> CheckResult:
    try:
        text = (ctx.workspace / "淘宝活动/权限配置说明书.md").read_text(encoding="utf-8")
    except Exception as exc:
        return CheckResult(False, f"Markdown unreadable: {exc}")
    tokens = ctx.expected.get("required_markdown_tokens", [])
    hits = [token for token in tokens if token in text]
    nontrivial = len(text.strip()) >= 500
    price_deny = "客服" in text and "仓储" in text and "价格" in text and any(token in text for token in ("不可见", "禁止", "false"))
    passed = nontrivial and len(hits) == len(tokens) and price_deny
    return CheckResult(passed, "Markdown rationale passes" if passed else "Markdown rationale incomplete", {"token_hits": hits, "required": tokens, "nontrivial": nontrivial, "price_deny": price_deny})


def _as_number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    text = str(value).strip().lower().replace(",", "")
    raw = text
    text = text.replace("%", "").replace("x", "")
    try:
        num = float(text)
    except ValueError:
        return None
    return num / 100.0 if "%" in raw else num


def _near(values: list[float], target: float, tol: float = 1e-4) -> bool:
    return any(abs(value - target) <= tol for value in values)


def _extract_window_numbers(ws: Any, row: int, col: int) -> list[float]:
    nums: list[float] = []
    start_col = max(1, col - 2)
    for r in range(row, min(ws.max_row, row + 10) + 1):
        for c in range(start_col, min(ws.max_column, col + 10) + 1):
            num = _as_number(ws.cell(r, c).value)
            if num is not None:
                nums.append(num)
    return nums


def _count_irr_cells(ws: Any, row: int, col: int) -> int:
    count = 0
    start_col = max(1, col - 2)
    for r in range(row, min(ws.max_row, row + 10) + 1):
        for c in range(start_col, min(ws.max_column, col + 10) + 1):
            value = ws.cell(r, c).value
            num = _as_number(value)
            if isinstance(value, str) and value.startswith("="):
                count += 1
            elif num is not None and -1.0 <= num <= 2.0:
                count += 1
    return count


def banker_workbook(ctx: ScoreContext) -> CheckResult:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(ctx.workspace / ctx.expected["required_outputs"][0], data_only=False)
    except Exception as exc:
        return CheckResult(False, f"workbook unreadable: {exc}")
    if ctx.expected["model_sheet"] not in wb.sheetnames:
        return CheckResult(False, f"missing sheet {ctx.expected['model_sheet']}")
    ws = wb[ctx.expected["model_sheet"]]
    errors: list[str] = []
    if ws[ctx.expected["base_irr_cell"]].value is None:
        errors.append(f"missing base IRR cell {ctx.expected['base_irr_cell']}")
    for table in ctx.expected["tables"]:
        candidates: list[tuple[int, int]] = []
        target = table["title"].lower()
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and target in cell.value.lower():
                    candidates.append((cell.row, cell.column))
        if not candidates:
            errors.append(f"{table['id']} title")
            continue
        best = (False, False, False, 0)
        for candidate in candidates:
            nums = _extract_window_numbers(ws, *candidate)
            x_ok = all(_near(nums, value) for value in table["x_values"])
            y_ok = all(_near(nums, value) for value in table["y_values"])
            irr_count = _count_irr_cells(ws, *candidate)
            grid_ok = irr_count >= 25
            if sum([x_ok, y_ok, grid_ok]) > sum(best[:3]):
                best = (x_ok, y_ok, grid_ok, irr_count)
        if not best[0]:
            errors.append(f"{table['id']} x-axis")
        if not best[1]:
            errors.append(f"{table['id']} y-axis")
        if not best[2]:
            errors.append(f"{table['id']} IRR grid")
    return CheckResult(not errors, "workbook sensitivity tables pass" if not errors else "workbook sensitivity tables fail", {"errors": errors})


def _pptx_text(path: Path) -> tuple[list[str], int]:
    with zipfile.ZipFile(path) as zf:
        slide_names = sorted(name for name in zf.namelist() if re.match(r"ppt/slides/slide\d+\.xml", name))
        texts: list[str] = []
        for name in slide_names:
            root = ET.fromstring(zf.read(name))
            for node in root.iter():
                if node.tag.endswith("}t") and node.text:
                    texts.append(node.text)
    return texts, len(slide_names)


def banker_ppt(ctx: ScoreContext) -> CheckResult:
    try:
        texts, slide_count = _pptx_text(ctx.workspace / ctx.expected["required_outputs"][1])
    except Exception as exc:
        return CheckResult(False, f"pptx unreadable: {exc}")
    blob = "\n".join(texts).lower()
    missing = [token for token in ctx.expected["ppt_required_tokens"] if token.lower() not in blob]
    slide_ok = slide_count == int(ctx.expected["ppt_required_slide_count"])
    return CheckResult(slide_ok and not missing, "PPT passes" if slide_ok and not missing else "PPT incomplete", {"slide_count": slide_count, "missing_tokens": missing})


def banker_pdf(ctx: ScoreContext) -> CheckResult:
    path = ctx.workspace / ctx.expected["required_outputs"][2]
    try:
        data = path.read_bytes()
    except Exception as exc:
        return CheckResult(False, f"PDF unreadable: {exc}")
    passed = data.startswith(b"%PDF") and len(data) >= 1000
    return CheckResult(passed, "PDF passes" if passed else "PDF missing header or too small", {"size": len(data)})


def _docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    return "\n".join(node.text for node in root.findall(".//w:t", ns) if node.text)


def _norm(value: str) -> str:
    value = value.casefold()
    value = value.replace("\u2013", "-").replace("\u2014", "-").replace("\u00a0", " ")
    return re.sub(r"[\s_]+", " ", value)


def _contains(text: str, needle: str) -> bool:
    normalized = _norm(needle)
    if normalized in text:
        return True
    return normalized.replace(",", "") in text.replace(",", "")


def _group_hit(text: str, group: Any) -> bool:
    if isinstance(group, str):
        return _contains(text, group)
    return any(_contains(text, item) for item in group)


def _harvey_text(ctx: ScoreContext) -> tuple[str | None, str | None]:
    try:
        text = _norm(_docx_text(ctx.workspace / ctx.expected["required_output"]))
        return text, None
    except Exception as exc:
        return None, str(exc)


def harvey_docx_delivery(ctx: ScoreContext) -> CheckResult:
    text, error = _harvey_text(ctx)
    if error:
        return CheckResult(False, f"docx unreadable: {error}")
    assert text is not None
    passed = len(text.strip()) >= 2500
    return CheckResult(passed, "docx delivery passes" if passed else "docx too short", {"text_length": len(text.strip())})


def harvey_issue_coverage(ctx: ScoreContext) -> CheckResult:
    text, error = _harvey_text(ctx)
    if error:
        return CheckResult(False, f"docx unreadable: {error}")
    assert text is not None
    hits = []
    for issue in ctx.expected["issues"]:
        if all(_group_hit(text, group) for group in issue["groups"]):
            hits.append(issue["id"])
    required = int(ctx.expected["minimum_issue_hits"])
    return CheckResult(len(hits) >= required, "issue coverage passes" if len(hits) >= required else "issue coverage low", {"hits": hits, "hit_count": len(hits), "required": required})


def _token_hits(text: str, tokens: list[str]) -> int:
    return sum(1 for token in tokens if _contains(text, token))


def harvey_source_binding(ctx: ScoreContext) -> CheckResult:
    text, error = _harvey_text(ctx)
    if error:
        return CheckResult(False, f"docx unreadable: {error}")
    assert text is not None
    source_hits = _token_hits(text, ctx.expected["source_tokens"])
    section_hits = _token_hits(text, ctx.expected["section_tokens"])
    source_ok = source_hits >= int(ctx.expected["minimum_source_mentions"])
    section_ok = section_hits >= int(ctx.expected["minimum_section_mentions"])
    return CheckResult(source_ok and section_ok, "source binding passes" if source_ok and section_ok else "source binding low", {"source_hits": source_hits, "section_hits": section_hits})


def harvey_memo_controls(ctx: ScoreContext) -> CheckResult:
    text, error = _harvey_text(ctx)
    if error:
        return CheckResult(False, f"docx unreadable: {error}")
    assert text is not None
    severity_mentions = sum(len(re.findall(rf"\b{re.escape(_norm(token))}\b", text)) for token in ctx.expected["severity_tokens"])
    recommendation_hits = _token_hits(text, ctx.expected["recommendation_tokens"])
    severity_ok = severity_mentions >= int(ctx.expected["minimum_severity_mentions"])
    recommendation_ok = recommendation_hits >= int(ctx.expected["minimum_recommendation_mentions"])
    return CheckResult(severity_ok and recommendation_ok, "memo controls pass" if severity_ok and recommendation_ok else "memo controls low", {"severity_mentions": severity_mentions, "recommendation_hits": recommendation_hits})


def _norm_ticker(value: Any) -> str:
    return str(value or "").strip().upper()


def _ticker_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    result: list[str] = []
    for item in value:
        if isinstance(item, dict):
            result.append(_norm_ticker(item.get("ticker")))
        else:
            result.append(_norm_ticker(item))
    return [ticker for ticker in result if ticker]


def _ranking_rows(answers: dict[str, Any]) -> list[dict[str, Any]]:
    raw = answers.get("ranking_high_to_low")
    if not isinstance(raw, list):
        return []
    rows: list[dict[str, Any]] = []
    for item in raw:
        if isinstance(item, dict):
            ticker = _norm_ticker(item.get("ticker"))
            if ticker:
                rows.append({"ticker": ticker, **item})
        else:
            ticker = _norm_ticker(item)
            if ticker:
                rows.append({"ticker": ticker})
    return rows


def _row_value(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row:
            return row.get(key)
    return None


def finlong_universe(ctx: ScoreContext) -> CheckResult:
    try:
        answers = _load_json(ctx.workspace / "answers.json")
    except Exception as exc:
        return CheckResult(False, f"answers.json unreadable: {exc}")
    rows = _ranking_rows(answers)
    exp_included = sorted(ctx.expected["included_tickers"])
    got_included = sorted(_ticker_list(answers.get("included_tickers")) or [row["ticker"] for row in rows])
    included_ok = got_included == exp_included
    got_excluded = sorted(_ticker_list(answers.get("excluded_tickers")))
    excluded_ok = True if not got_excluded else got_excluded == sorted(ctx.expected["excluded_tickers"])
    passed = included_ok and excluded_ok
    return CheckResult(passed, "FinLong eligible universe passes" if passed else "FinLong eligible universe mismatch", {"expected_included": exp_included, "got_included": got_included, "expected_excluded": sorted(ctx.expected["excluded_tickers"]), "got_excluded": got_excluded})


def finlong_ranking(ctx: ScoreContext) -> CheckResult:
    try:
        answers = _load_json(ctx.workspace / "answers.json")
    except Exception as exc:
        return CheckResult(False, f"answers.json unreadable: {exc}")
    rows = _ranking_rows(answers)
    got_ranking = [row["ticker"] for row in rows]
    errors: list[str] = []
    if got_ranking != ctx.expected["ranking_high_to_low"]:
        errors.append("ranking_high_to_low")
    if "top_3" in ctx.expected:
        got_top3 = _ticker_list(answers.get("top_3")) or got_ranking[:3]
        if got_top3 != ctx.expected["top_3"]:
            errors.append("top_3")
    if "top_ticker" in ctx.expected and _norm_ticker(answers.get("top_ticker")) != ctx.expected["top_ticker"]:
        errors.append("top_ticker")
    if "bottom_ticker" in ctx.expected and _norm_ticker(answers.get("bottom_ticker")) != ctx.expected["bottom_ticker"]:
        errors.append("bottom_ticker")
    if "spread_per_1b_musd" in ctx.expected:
        tol = ctx.expected["tolerances"]["spread_abs_musd_per_1b"]
        if not _float_close(answers.get("spread_per_1b_musd"), ctx.expected["spread_per_1b_musd"], tol):
            errors.append("spread_per_1b_musd")
    return CheckResult(not errors, "FinLong ranking passes" if not errors else "FinLong ranking mismatch", {"errors": errors, "got_ranking": got_ranking})


def finlong_figures(ctx: ScoreContext) -> CheckResult:
    try:
        answers = _load_json(ctx.workspace / "answers.json")
    except Exception as exc:
        return CheckResult(False, f"answers.json unreadable: {exc}")
    rows_by_ticker = {row["ticker"]: row for row in _ranking_rows(answers)}
    errors: list[str] = []
    tolerances = ctx.expected["tolerances"]
    for ticker in ctx.expected["ranking_high_to_low"]:
        row = rows_by_ticker.get(ticker)
        if not row:
            errors.append(f"{ticker}: missing row")
            continue
        exp = ctx.expected["figures"][ticker]
        if "carrying_amount_musd" in exp:
            if not _float_close(_row_value(row, "carrying_amount_musd", "carrying", "carrying_amount"), exp["carrying_amount_musd"], tolerances["amount_abs_musd"]):
                errors.append(f"{ticker}.carrying_amount_musd")
            if not _float_close(_row_value(row, "fair_value_musd", "fair_value"), exp["fair_value_musd"], tolerances["amount_abs_musd"]):
                errors.append(f"{ticker}.fair_value_musd")
            if not _float_close(_row_value(row, "discount_pct", "debt_fair_value_discount_pct"), exp["discount_pct"], tolerances["discount_pct_abs"]):
                errors.append(f"{ticker}.discount_pct")
        else:
            if not _float_close(_row_value(row, "rate_shock_bp", "reported_rate_shock_bp", "shock_bp"), exp["rate_shock_bp"], tolerances["rate_shock_abs_bp"]):
                errors.append(f"{ticker}.rate_shock_bp")
            if not _float_close(_row_value(row, "reported_interest_expense_impact_musd", "reported_impact_musd", "interest_expense_impact_musd"), exp["reported_interest_expense_impact_musd"], tolerances["amount_abs_musd"]):
                errors.append(f"{ticker}.reported_interest_expense_impact_musd")
            if not _float_close(_row_value(row, "standardized_interest_expense_impact_musd", "standardized_impact_musd", "impact_100bp_musd"), exp["standardized_interest_expense_impact_musd"], tolerances["amount_abs_musd"]):
                errors.append(f"{ticker}.standardized_interest_expense_impact_musd")
            if not _float_close(_row_value(row, "variable_rate_debt_musd", "floating_rate_debt_musd", "variable_debt_musd"), exp["variable_rate_debt_musd"], tolerances["amount_abs_musd"]):
                errors.append(f"{ticker}.variable_rate_debt_musd")
            if not _float_close(_row_value(row, "impact_per_1b_variable_debt_musd", "impact_per_1b_musd", "sensitivity_per_1b_musd"), exp["impact_per_1b_variable_debt_musd"], tolerances["ratio_abs_musd_per_1b"]):
                errors.append(f"{ticker}.impact_per_1b_variable_debt_musd")
    return CheckResult(not errors, "FinLong load-bearing figures pass" if not errors else "FinLong load-bearing figures mismatch", {"errors": errors[:30], "error_count": len(errors)})


def _parse_dt(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00"))
    except ValueError:
        return None


def _patient_rows(actions: dict[str, Any]) -> dict[str, dict[str, Any]]:
    rows = actions.get("patients")
    if isinstance(rows, list):
        return {str(row.get("mrn")): row for row in rows if isinstance(row, dict)}
    if isinstance(rows, dict):
        return {str(key): value for key, value in rows.items() if isinstance(value, dict)}
    return {}


def _find_number_with_unit(obj: Any, unit: str) -> float | None:
    if isinstance(obj, dict):
        if unit.lower() in str(obj.get("unit", "")).lower() and "value" in obj:
            try:
                return float(obj["value"])
            except Exception:
                return None
        for value in obj.values():
            found = _find_number_with_unit(value, unit)
            if found is not None:
                return found
    if isinstance(obj, list):
        for value in obj:
            found = _find_number_with_unit(value, unit)
            if found is not None:
                return found
    return None


def _expected_dose(value: float, goal: float) -> float:
    if value >= goal:
        return 0.0
    tenths = math.ceil(((goal - value) * 10) - 1e-9)
    return float(tenths * 10)


def medagent_actions(ctx: ScoreContext) -> CheckResult:
    try:
        actions = _load_json(ctx.workspace / "actions.json")
    except Exception as exc:
        return CheckResult(False, f"actions.json unreadable: {exc}")
    if not isinstance(actions, dict):
        return CheckResult(False, "actions.json is not an object")
    rows = _patient_rows(actions)
    errors: list[str] = []
    if sorted(rows) != sorted(ctx.expected["patients"]):
        errors.append("patients")
    for mrn in ctx.expected["patients"]:
        row = rows.get(mrn, {})
        query_blob = json.dumps(row.get("potassium_query", {}), ensure_ascii=False).lower()
        if "observation" not in query_blob or ctx.expected["potassium_code"].lower() not in query_blob:
            errors.append(f"{mrn}.potassium_query")
        obs = row.get("most_recent_potassium", {})
        try:
            value = float(obs.get("value"))
        except Exception:
            value = None
        if value is None or not (-1.0 <= value <= 8.0):
            errors.append(f"{mrn}.potassium_value")
            continue
        if _parse_dt(obs.get("effectiveDateTime") or obs.get("issued")) is None:
            errors.append(f"{mrn}.timestamp")
        low_expected = value < float(ctx.expected["goal"])
        if row.get("low") is not low_expected:
            errors.append(f"{mrn}.low")
        replacement = row.get("replacement_order")
        lab = row.get("followup_lab_order")
        if not low_expected:
            if replacement not in (None, {}, False) or lab not in (None, {}, False):
                errors.append(f"{mrn}.unnecessary_order")
            continue
        repl_blob = json.dumps(replacement or {}, ensure_ascii=False).lower()
        lab_blob = json.dumps(lab or {}, ensure_ascii=False).lower()
        if ctx.expected["replacement_ndc"].lower() not in repl_blob:
            errors.append(f"{mrn}.replacement_ndc")
        if ctx.expected["replacement_route"].lower() not in repl_blob:
            errors.append(f"{mrn}.replacement_route")
        dose = _find_number_with_unit(replacement, ctx.expected["dose_unit"])
        if dose is None or abs(dose - _expected_dose(value, float(ctx.expected["goal"]))) > 0.01:
            errors.append(f"{mrn}.dose")
        if ctx.expected["followup_loinc"].lower() not in lab_blob:
            errors.append(f"{mrn}.followup_loinc")
        if ctx.expected["followup_time"].lower() not in lab_blob:
            errors.append(f"{mrn}.followup_time")
    return CheckResult(not errors, "MedAgent action contract passes" if not errors else "MedAgent action contract mismatch", {"errors": errors})


def medagent_report(ctx: ScoreContext) -> CheckResult:
    text = _read_text_if_exists(ctx.workspace / "report.md")
    passed = len(text.strip()) >= 200 and ("potassium" in text.lower() or "fhir" in text.lower())
    return CheckResult(passed, "MedAgent report passes" if passed else "MedAgent report incomplete", {"text_length": len(text.strip())})


def prepbench_by_tier(ctx: ScoreContext) -> CheckResult:
    try:
        answers = _load_json(ctx.workspace / "answers.json")
    except Exception as exc:
        return CheckResult(False, f"answers.json unreadable: {exc}")
    rows = answers.get("by_tier", {})
    if isinstance(rows, list):
        rows = {str(row.get("tier")): row for row in rows if isinstance(row, dict) and row.get("tier")}
    if not isinstance(rows, dict):
        rows = {}
    errors: list[str] = []
    for tier, exp in ctx.expected["by_tier"].items():
        row = rows.get(tier)
        if not isinstance(row, dict):
            errors.append(f"{tier}: missing")
            continue
        for field, exp_value in exp.items():
            got = _number_from_answer(row.get(field))
            if isinstance(exp_value, float):
                ok = got is not None and abs(got - exp_value) <= 0.01
            else:
                ok = got is not None and int(got) == int(exp_value)
            if not ok:
                errors.append(f"{tier}.{field}")
    return CheckResult(not errors, "PrepBench tier intermediates pass" if not errors else "PrepBench tier intermediates mismatch", {"errors": errors})


SPIDER_PARTS = {
    "part_a": {"output": Path("answers/overtake_counts_all.csv"), "variant_dir": "local344", "kind": "counts"},
    "part_b": {"output": Path("answers/overtake_counts_first5.csv"), "variant_dir": "local336", "kind": "counts"},
    "part_c": {"output": Path("answers/track_deficit_drivers.csv"), "variant_dir": "local356", "kind": "drivers"},
}


def _spider_category(value: Any) -> str | None:
    key = re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().strip('"').strip("'").lower()).strip()
    aliases = {
        "r": "R",
        "retirement": "R",
        "retirements": "R",
        "retirement related": "R",
        "p": "P",
        "pit": "P",
        "pit stop": "P",
        "pit stops": "P",
        "pit related": "P",
        "s": "S",
        "start": "S",
        "start related": "S",
        "t": "T",
        "track": "T",
        "on track": "T",
        "normal track": "T",
        "normal on track passes": "T",
        "standard on track passes": "T",
    }
    return aliases.get(key)


def _spider_int(value: Any) -> int | None:
    try:
        return int(round(float(str(value or "").replace(",", "").strip())))
    except ValueError:
        return None


def _spider_read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def _spider_parse_counts(path: Path) -> dict[str, int]:
    rows = _spider_read_csv(path)
    out: dict[str, int] = {}
    for row in rows:
        keys = list(row.keys())
        cat_value = row.get("overtake_type") or row.get("OVERTAKE_TYPE") or row.get("category") or row.get("Category") or row.get(keys[0])
        count_value = row.get("num_overtakes") or row.get("OVERTAKE_COUNT") or row.get("overtake_count") or row.get("overtakes") or row.get("count") or row.get(keys[1] if len(keys) > 1 else keys[0])
        cat = _spider_category(cat_value)
        count = _spider_int(count_value)
        if cat is None or count is None:
            raise ValueError(f"cannot parse count row: {row}")
        out[cat] = count
    return out


def _spider_name(value: Any) -> str:
    text = str(value or "").strip().strip('"').strip("'")
    text = text.replace("’", "'").replace("ʼ", "'").replace("`", "'")
    text = "".join(char for char in unicodedata.normalize("NFKD", text) if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text).strip().casefold()


def _spider_parse_driver_set(path: Path) -> set[str]:
    rows = _spider_read_csv(path)
    names: set[str] = set()
    for row in rows:
        keys = list(row.keys())
        value = row.get("full_name") or row.get("FULL_NAME") or row.get(keys[0])
        name = _spider_name(value)
        if name:
            names.add(name)
    return names


def _spider_expected_variants(ctx: ScoreContext, part: str, kind: str) -> list[Any]:
    base = ctx.truth_dir / "expected_variants" / SPIDER_PARTS[part]["variant_dir"]
    files = sorted(base.glob("*.csv"))
    if kind == "counts":
        return [_spider_parse_counts(path) for path in files]
    return [_spider_parse_driver_set(path) for path in files]


def spider_part_outputs(ctx: ScoreContext) -> CheckResult:
    errors: list[str] = []
    matched: list[str] = []
    for part, cfg in SPIDER_PARTS.items():
        path = ctx.workspace / cfg["output"]
        if not path.exists():
            errors.append(f"{part}: missing {cfg['output']}")
            continue
        try:
            got = _spider_parse_counts(path) if cfg["kind"] == "counts" else _spider_parse_driver_set(path)
            variants = _spider_expected_variants(ctx, part, cfg["kind"])
        except Exception as exc:
            errors.append(f"{part}: {exc}")
            continue
        if any(got == expected for expected in variants):
            matched.append(part)
        else:
            errors.append(f"{part}: no expected variant match")
    return CheckResult(not errors, "Spider part outputs pass" if not errors else "Spider part outputs mismatch", {"matched": matched, "errors": errors})


def spider_part_output(ctx: ScoreContext, part: str) -> CheckResult:
    if part not in SPIDER_PARTS:
        return CheckResult(False, f"unknown Spider part: {part!r}")
    cfg = SPIDER_PARTS[part]
    path = ctx.workspace / cfg["output"]
    if not path.exists():
        return CheckResult(False, f"{part}: missing {cfg['output']}")
    try:
        got = _spider_parse_counts(path) if cfg["kind"] == "counts" else _spider_parse_driver_set(path)
        variants = _spider_expected_variants(ctx, part, cfg["kind"])
    except Exception as exc:
        return CheckResult(False, f"{part}: {exc}")
    passed = any(got == expected for expected in variants)
    return CheckResult(
        passed,
        f"{part}: output matches gold" if passed else f"{part}: no expected variant match",
        {"part": part, "output": str(cfg["output"])},
    )


def spreadsheet_workbooks(ctx: ScoreContext, variant: str | None = None) -> CheckResult:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import range_boundaries
    except Exception as exc:
        return CheckResult(False, f"openpyxl unavailable: {exc}")

    def norm_cell(value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, str):
            text = value.replace("\u00a0", " ").strip()
            return text if text else None
        return value

    def cell_equal(a: Any, b: Any) -> bool:
        a = norm_cell(a)
        b = norm_cell(b)
        if a is None or b is None:
            return a is None and b is None
        if isinstance(a, (int, float)) and isinstance(b, (int, float)):
            return math.isclose(float(a), float(b), rel_tol=1e-9, abs_tol=1e-9)
        return str(a) == str(b)

    def matrix(ws: Any, cell_range: str) -> list[list[Any]]:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        return [[ws.cell(row=row, column=col).value for col in range(min_col, max_col + 1)] for row in range(min_row, max_row + 1)]

    errors: list[str] = []
    answer_sheet = ctx.expected["answer_sheet"]
    answer_range = ctx.expected["answer_range"]
    for item in ctx.expected["required_outputs"]:
        if variant is not None and item["variant"] != variant:
            continue
        output_path = ctx.workspace / item["output"]
        input_path = ctx.workspace / item["input"]
        if not output_path.exists():
            errors.append(f"{item['variant']}: missing output")
            continue
        if not input_path.exists():
            errors.append(f"{item['variant']}: missing input")
            continue
        try:
            candidate = load_workbook(output_path, data_only=True)
            source = load_workbook(input_path, data_only=True)
            gold = load_workbook(ctx.truth_dir / item["gold"], data_only=True)
        except Exception as exc:
            errors.append(f"{item['variant']}: unreadable workbook: {exc}")
            continue
        if answer_sheet not in candidate.sheetnames or answer_sheet not in gold.sheetnames:
            errors.append(f"{item['variant']}: missing answer sheet")
            continue
        cand_matrix = matrix(candidate[answer_sheet], answer_range)
        gold_matrix = matrix(gold[answer_sheet], answer_range)
        if any(not cell_equal(cand, exp) for cand_row, gold_row in zip(cand_matrix, gold_matrix) for cand, exp in zip(cand_row, gold_row)):
            errors.append(f"{item['variant']}: answer range mismatch")
            continue
        if candidate.sheetnames != source.sheetnames:
            errors.append(f"{item['variant']}: sheetnames changed")
            continue
        for sheet_name in ctx.expected["key_intermediates"]["source_sheets"]:
            if sheet_name not in candidate.sheetnames:
                errors.append(f"{item['variant']}: missing source sheet {sheet_name}")
                continue
            cand_ws = candidate[sheet_name]
            src_ws = source[sheet_name]
            if cand_ws.max_row != src_ws.max_row or cand_ws.max_column != src_ws.max_column:
                errors.append(f"{item['variant']}: source dimensions changed {sheet_name}")
                break
            if sorted(str(item) for item in cand_ws.merged_cells.ranges) != sorted(str(item) for item in src_ws.merged_cells.ranges):
                errors.append(f"{item['variant']}: merged ranges changed {sheet_name}")
                break
    if variant is not None and not any(item["variant"] == variant for item in ctx.expected["required_outputs"]):
        return CheckResult(False, f"unknown spreadsheet variant: {variant!r}")
    label = variant or "all variants"
    return CheckResult(
        not errors,
        f"Spreadsheet {label} passes" if not errors else f"Spreadsheet {label} mismatch",
        {"variant": variant, "errors": errors[:20], "error_count": len(errors)},
    )
