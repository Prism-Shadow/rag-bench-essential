#!/usr/bin/env python3
from __future__ import annotations

import json
import os
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


def truth_dir() -> Path:
    return Path(os.environ.get("BENCH_TRUTH_DIR", Path(__file__).resolve().parent))


def load_expected() -> dict:
    return json.loads((truth_dir() / "expected.json").read_text(encoding="utf-8"))


def as_number(value) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return None
    text = str(value).strip().lower().replace(",", "")
    text = text.replace("%", "").replace("x", "")
    try:
        num = float(text)
    except ValueError:
        return None
    if "%" in str(value):
        return num / 100.0
    return num


def near(values: list[float], target: float, tol: float = 1e-4) -> bool:
    return any(abs(value - target) <= tol for value in values)


def extract_window_numbers(ws, row: int, col: int) -> list[float]:
    nums: list[float] = []
    start_col = max(1, col - 2)
    for r in range(row, min(ws.max_row, row + 10) + 1):
        for c in range(start_col, min(ws.max_column, col + 10) + 1):
            n = as_number(ws.cell(r, c).value)
            if n is not None:
                nums.append(n)
    return nums


def count_irr_cells(ws, row: int, col: int) -> int:
    count = 0
    start_col = max(1, col - 2)
    for r in range(row, min(ws.max_row, row + 10) + 1):
        for c in range(start_col, min(ws.max_column, col + 10) + 1):
            value = ws.cell(r, c).value
            if isinstance(value, str) and value.startswith("="):
                count += 1
            else:
                n = as_number(value)
                if n is not None and -1.0 <= n <= 2.0:
                    count += 1
    return count


def find_title_cells(ws, title: str) -> list[tuple[int, int]]:
    target = title.lower()
    found: list[tuple[int, int]] = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and target in value.lower():
                found.append((cell.row, cell.column))
    return found


def score_workbook(expected: dict) -> bool:
    path = Path(expected["required_outputs"][0])
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as exc:
        print(f"  [MISS] workbook unreadable: {exc}")
        return False
    if expected["model_sheet"] not in wb.sheetnames:
        print(f"  [MISS] missing sheet {expected['model_sheet']}")
        return False
    ws = wb[expected["model_sheet"]]
    base_formula_ok = ws[expected["base_irr_cell"]].value is not None
    print(f"  [{'OK ' if base_formula_ok else 'MISS'}] base IRR cell {expected['base_irr_cell']} present")
    ok = base_formula_ok
    for table in expected["tables"]:
        candidates = find_title_cells(ws, table["title"])
        if not candidates:
            print(f"  [MISS] table title not found: {table['title']}")
            ok = False
            continue
        best = None
        for found in candidates:
            nums = extract_window_numbers(ws, *found)
            x_ok_candidate = all(near(nums, value) for value in table["x_values"])
            y_ok_candidate = all(near(nums, value) for value in table["y_values"])
            irr_count_candidate = count_irr_cells(ws, *found)
            grid_ok_candidate = irr_count_candidate >= 25
            score = sum([x_ok_candidate, y_ok_candidate, grid_ok_candidate])
            if best is None or score > best[0]:
                best = (score, x_ok_candidate, y_ok_candidate, grid_ok_candidate, irr_count_candidate)
        assert best is not None
        _score, x_ok, y_ok, grid_ok, irr_count = best
        print(f"  [{'OK ' if x_ok else 'MISS'}] {table['id']} x-axis values")
        print(f"  [{'OK ' if y_ok else 'MISS'}] {table['id']} y-axis values")
        print(f"  [{'OK ' if grid_ok else 'MISS'}] {table['id']} has >=25 IRR/formula cells (found {irr_count})")
        ok = ok and x_ok and y_ok and grid_ok
    return ok


def pptx_text(path: Path) -> tuple[list[str], int]:
    with zipfile.ZipFile(path) as zf:
        slide_names = sorted(name for name in zf.namelist() if re.match(r"ppt/slides/slide\d+\.xml", name))
        texts: list[str] = []
        for name in slide_names:
            root = ET.fromstring(zf.read(name))
            for node in root.iter():
                if node.tag.endswith("}t") and node.text:
                    texts.append(node.text)
    return texts, len(slide_names)


def score_ppt(expected: dict) -> bool:
    path = Path(expected["required_outputs"][1])
    try:
        texts, slide_count = pptx_text(path)
    except Exception as exc:
        print(f"  [MISS] pptx unreadable: {exc}")
        return False
    blob = "\n".join(texts).lower()
    slide_ok = slide_count == int(expected["ppt_required_slide_count"])
    print(f"  [{'OK ' if slide_ok else 'MISS'}] slide count={slide_count}")
    tokens_ok = True
    for token in expected["ppt_required_tokens"]:
        hit = token.lower() in blob
        print(f"  [{'OK ' if hit else 'MISS'}] PPT token: {token}")
        tokens_ok = tokens_ok and hit
    return slide_ok and tokens_ok


def score_pdf(expected: dict) -> bool:
    path = Path(expected["required_outputs"][2])
    try:
        data = path.read_bytes()
    except Exception as exc:
        print(f"  [MISS] PDF unreadable: {exc}")
        return False
    ok = data.startswith(b"%PDF") and len(data) >= 1000
    print(f"  [{'OK ' if ok else 'MISS'}] PDF exists, has PDF header, and is non-trivial")
    return ok


def main() -> int:
    expected = load_expected()
    missing = [name for name in expected["required_outputs"] if not Path(name).exists()]
    if missing:
        print("FAIL: required output files missing:")
        for name in missing:
            print(f"  - {name}")
        return 2
    print("== D1/D3 Excel sensitivity tables ==")
    workbook_ok = score_workbook(expected)
    print("== D4 PowerPoint deck ==")
    ppt_ok = score_ppt(expected)
    print("== D4 PDF export ==")
    pdf_ok = score_pdf(expected)
    if workbook_ok and ppt_ok and pdf_ok:
        print("\nRESULT: PASS - Excel/PPT/PDF sensitivity deliverables satisfy the deterministic gates.")
        return 0
    print("\nRESULT: FAIL - sensitivity table structure or deliverables are incomplete.")
    return 1


if __name__ == "__main__":
    sys.exit(main())
