#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
from pathlib import Path

from openpyxl import load_workbook


SOURCE_SHEETS = ["lap-1", "lap-2", "lap-9"]
SLOTS_PER_ROW = 10
FIELDS_PER_BLOCK = 12
ANSWER_ROWS = 8
ANSWER_COLS = 121  # A:DQ

# Coordinate order from task.md. Offsets are relative to the block start row.
FIELD_OFFSETS = [
    (2, 3),   # C3
    (3, 3),   # C4
    (4, 3),   # C5
    (6, 3),   # C7
    (7, 3),   # C8
    (10, 3),  # C11
    (11, 3),  # C12
    (8, 5),   # E9
    (10, 5),  # E11
    (11, 5),  # E12
    (14, 3),  # C15
    (20, 3),  # C21
]


def truth_dir() -> Path:
    return Path(os.environ.get("BENCH_TRUTH_DIR", Path(__file__).resolve().parent))


def find_blocks(ws) -> list[int]:
    starts: list[int] = []
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value is None:
            continue
        text = str(value)
        if "Calculate" in text or "Perhitungan Ke" in text:
            starts.append(row)
    return starts


def extract_block(ws, start_row: int) -> list[object]:
    return [
        ws.cell(row=start_row + row_offset, column=col_idx).value
        for row_offset, col_idx in FIELD_OFFSETS
    ]


def clear_answer_range(ws) -> None:
    for row in range(1, ANSWER_ROWS + 1):
        for col in range(1, ANSWER_COLS + 1):
            ws.cell(row=row, column=col).value = None


def complete_workbook(input_path: Path, output_path: Path) -> None:
    wb = load_workbook(input_path)
    ws_dest = wb["destination"]

    clear_answer_range(ws_dest)

    for slot_idx in range(SLOTS_PER_ROW):
        start_col = 2 + slot_idx * FIELDS_PER_BLOCK
        for field_idx in range(FIELDS_PER_BLOCK):
            ws_dest.cell(row=1, column=start_col + field_idx).value = field_idx + 1

    dest_row = 2
    output_index = 1
    for sheet_name in SOURCE_SHEETS:
        ws_source = wb[sheet_name]
        starts = find_blocks(ws_source)
        for chunk_start in range(0, len(starts), SLOTS_PER_ROW):
            if dest_row > ANSWER_ROWS:
                raise ValueError(f"not enough destination rows for {input_path}")
            chunk = starts[chunk_start : chunk_start + SLOTS_PER_ROW]
            ws_dest.cell(row=dest_row, column=1).value = output_index
            for slot_idx, block_start in enumerate(chunk):
                values = extract_block(ws_source, block_start)
                start_col = 2 + slot_idx * FIELDS_PER_BLOCK
                for field_idx, value in enumerate(values):
                    ws_dest.cell(row=dest_row, column=start_col + field_idx).value = value
            dest_row += 1
            output_index += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--write-gold",
        action="store_true",
        help="write truth/expected_variants gold workbooks instead of workspace outputs",
    )
    args = parser.parse_args()

    truth = truth_dir()
    spec = json.loads((truth / "expected.json").read_text())
    case_root = truth.parent
    workspace = Path.cwd()

    for item in spec["required_outputs"]:
        workspace_input = workspace / item["input"]
        input_path = workspace_input if workspace_input.exists() else case_root / item["input"]
        output_path = truth / item["gold"] if args.write_gold else workspace / item["output"]
        complete_workbook(input_path, output_path)
        print(f"wrote {output_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
