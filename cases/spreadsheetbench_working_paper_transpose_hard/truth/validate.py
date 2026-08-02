#!/usr/bin/env python3
from __future__ import annotations

import json
import math
import os
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
except Exception as exc:  # pragma: no cover - dependency failure path
    print(f"VALIDATION ERROR: openpyxl is required: {exc}")
    sys.exit(2)


def truth_dir() -> Path:
    return Path(os.environ.get("BENCH_TRUTH_DIR", Path(__file__).resolve().parent))


def norm(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("\u00a0", " ").strip()
        return text if text else None
    return value


def equal(a: Any, b: Any) -> bool:
    a = norm(a)
    b = norm(b)
    if a is None or b is None:
        return a is None and b is None
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return math.isclose(float(a), float(b), rel_tol=1e-9, abs_tol=1e-9)
    return str(a) == str(b)


def matrix(ws, cell_range: str) -> list[list[Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    return [
        [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
        for r in range(min_row, max_row + 1)
    ]


def merged_ranges(ws) -> list[str]:
    return sorted(str(item) for item in ws.merged_cells.ranges)


def compare_source_sheets(candidate_wb, source_wb, source_sheets: list[str]) -> list[str]:
    mismatches: list[str] = []
    if candidate_wb.sheetnames != source_wb.sheetnames:
        mismatches.append(f"sheetnames changed: got {candidate_wb.sheetnames!r}, expected {source_wb.sheetnames!r}")
        return mismatches

    for sheet_name in source_sheets:
        if sheet_name not in candidate_wb.sheetnames:
            mismatches.append(f"missing preserved source sheet {sheet_name!r}")
            continue
        cand_ws = candidate_wb[sheet_name]
        src_ws = source_wb[sheet_name]
        if cand_ws.max_row != src_ws.max_row or cand_ws.max_column != src_ws.max_column:
            mismatches.append(
                f"{sheet_name}: dimensions changed from "
                f"{src_ws.max_row}x{src_ws.max_column} to {cand_ws.max_row}x{cand_ws.max_column}"
            )
        if merged_ranges(cand_ws) != merged_ranges(src_ws):
            mismatches.append(
                f"{sheet_name}: merged ranges changed "
                f"({len(merged_ranges(cand_ws))} vs {len(merged_ranges(src_ws))})"
            )

        max_row = max(cand_ws.max_row, src_ws.max_row)
        max_col = max(cand_ws.max_column, src_ws.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cand = cand_ws.cell(row=row, column=col).value
                src = src_ws.cell(row=row, column=col).value
                if not equal(cand, src):
                    cell = src_ws.cell(row=row, column=col).coordinate
                    mismatches.append(
                        f"{sheet_name}!{cell}: got {cand!r}, expected preserved value {src!r}"
                    )
                    if len(mismatches) >= 20:
                        return mismatches

    return mismatches


def validate_one(workspace: Path, truth: Path, spec: dict[str, Any], item: dict[str, str]) -> tuple[bool, str]:
    output_path = workspace / item["output"]
    if not output_path.exists():
        return False, f"{item['variant']}: missing required output {item['output']}"

    input_path = workspace / item["input"]
    if not input_path.exists():
        return False, f"{item['variant']}: missing public input workbook {item['input']}"

    try:
        candidate_wb = load_workbook(output_path, data_only=True)
    except Exception as exc:
        return False, f"{item['variant']}: unreadable workbook {item['output']}: {exc}"

    try:
        source_wb = load_workbook(input_path, data_only=True)
    except Exception as exc:
        return False, f"{item['variant']}: unreadable public input workbook {item['input']}: {exc}"

    try:
        gold_wb = load_workbook(truth / item["gold"], data_only=True)
    except Exception as exc:
        return False, f"{item['variant']}: unreadable gold workbook: {exc}"

    sheet = spec["answer_sheet"]
    if sheet not in candidate_wb.sheetnames:
        return False, f"{item['variant']}: missing sheet {sheet!r}"
    if sheet not in gold_wb.sheetnames:
        return False, f"{item['variant']}: internal gold missing sheet {sheet!r}"

    cell_range = spec["answer_range"]
    candidate = matrix(candidate_wb[sheet], cell_range)
    gold = matrix(gold_wb[sheet], cell_range)

    mismatches: list[str] = []
    min_col, min_row, _, _ = range_boundaries(cell_range)
    for r_idx, (cand_row, gold_row) in enumerate(zip(candidate, gold), start=min_row):
        for c_offset, (cand, exp) in enumerate(zip(cand_row, gold_row), start=0):
            if not equal(cand, exp):
                col_idx = min_col + c_offset
                cell = candidate_wb[sheet].cell(row=r_idx, column=col_idx).coordinate
                mismatches.append(f"{cell}: got {cand!r}, expected {exp!r}")
                if len(mismatches) >= 20:
                    break
        if len(mismatches) >= 20:
            break

    if mismatches:
        return False, f"{item['variant']}: {len(mismatches)}+ mismatches in {sheet}!{cell_range}; first: " + "; ".join(mismatches)

    source_sheets = spec["key_intermediates"]["source_sheets"]
    source_mismatches = compare_source_sheets(candidate_wb, source_wb, source_sheets)
    if source_mismatches:
        return False, f"{item['variant']}: source workbook structure/value mismatch; first: " + "; ".join(source_mismatches[:20])

    return True, f"{item['variant']}: PASS {item['output']} matches {sheet}!{cell_range} and preserves source sheets"


def main() -> int:
    workspace = Path.cwd()
    truth = truth_dir()
    spec_path = truth / "expected.json"
    if not spec_path.exists():
        print(f"VALIDATION ERROR: missing expected.json at {spec_path}")
        return 2

    spec = json.loads(spec_path.read_text())
    messages: list[str] = []
    failures: list[str] = []
    missing = False

    for item in spec["required_outputs"]:
        ok, msg = validate_one(workspace, truth, spec, item)
        messages.append(msg)
        if not ok:
            failures.append(msg)
            if "missing required output" in msg or "unreadable" in msg:
                missing = True

    for msg in messages:
        print(msg)

    if failures:
        print(f"RESULT: FAIL ({len(failures)} / {len(spec['required_outputs'])} variants failed)")
        return 2 if missing else 1

    print(f"RESULT: PASS ({len(spec['required_outputs'])} variants matched)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
